import os
import sqlite3
import openpyxl

import mysql.connector
from mysql.connector import Error


def export_to_sqlite():
    """Экспорт данных из xlsx в sqlite"""

    connection = mysql.connector.connect(
        host="localhost",
        user="alisa",
        password="alisa24462",
        database="demo"
    )
    print("Connection to MySQL DB successful")
    cursor = connection.cursor()

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('Товар_import/Товар_import_Спортивные товары.xlsx', data_only=True)
    sheet = file_to_read['Лист1']

    # Цикл по строкам начиная со второй (в первой заголовки)
    for row in range(2, 32):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 12 ( 13 не включая)
        for col in range(1,13):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            print(value)
            # Список который мы потом будем добавлять
            data.append(value)

    # 3. Запись в базу и закрытие соединения
        if data[11] == None:
            data[11] ='xxxxx'
        print('row, col', row, col)
        print(data[0], data[1], data[10], data[7], data[11], data[5], data[3], data[4], data[9], data[8])
        # Вставка данных в поля таблицы
        val = (data[0], data[1], data[10], data[7], data[11], data[5], data[3], data[4], data[9], data[8])
        sql = ("INSERT INTO Product(ProductArticleNumber, ProductName, ProductDescription, ProductCategory, "
               "ProductPhoto, ProductManufacturer, ProductCost, ProductDiscountAmount, ProductQuantityInStock, "
               "ProductStatus) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)")
        cursor.execute(sql,val)

    # сохраняем изменения
    connection.commit()
    # закрытие соединения
    connection.close()


def user_export():
    connection = mysql.connector.connect(
        host="localhost",
        user="alisa",
        password="alisa24462",
        database="demo"
    )
    print("Connection to MySQL DB successful")
    cursor = connection.cursor()

    file_to_read = openpyxl.load_workbook('user_import.xlsx', data_only=True)
    sheet = file_to_read['Лист1']

    for row in range(2, 12):
        data = []
        for col in range(1, 5):
            value = sheet.cell(row, col).value
            print(value)
            data.append(value)

        sql_r = ("SELECT RoleID FROM Role WHERE RoleName = %s")
        val_r = (data[0],)
        cursor.execute(sql_r, val_r)
        fin = cursor.fetchall()[0][0]
        print(fin)
        fio = data[1].split()
        val = (fio[0], fio[1], fio[2], data[2], data[3], fin)
        print(val)

        sql = ("INSERT INTO User(UserSurname, UserName, UserPatronymic, "
               "UserLogin, UserPassword, UserRole) VALUES (%s,%s,%s,%s,%s,%s)")
        cursor.execute(sql, val)



    # # сохраняем изменения
    connection.commit()
    # # закрытие соединения
    connection.close()



def role_export():
    connection = mysql.connector.connect(
        host="localhost",
        user="alisa",
        password="alisa24462",
        database="demo"
    )
    print("Connection to MySQL DB successful")
    cursor = connection.cursor()

    data = ['Администратор', 'Менеджер', 'Клиент']
    for i in data:
        val = (i,)
        sql = ("INSERT INTO Role(RoleName) VALUES (%s)")
        cursor.execute(sql, val)
    # # сохраняем изменения
    connection.commit()
    # # закрытие соединения
    connection.close()


def clear_base():
    '''Очистка базы sqlite'''

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'auto.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM cars")
    connect.commit()
    connect.close()


# export_to_sqlite()
# user_export()
# role_export()
